import json
import datetime
from main import app, db
from common.other import *
from common.loader import pdf_loader
from flask import request, send_file
from models.basic import *
from models.sprint import Sprint
from models.nodes import Nodes
from models.llm_prompt import Prompt
from views.llm import llm_common_ask
from models.test_review import CaseReview
from utils.MSUtils import MSUtils
from views.login import require_login


@app.route('/api/case/extract_test_point', methods=["POST"])
def case_extract_test_point():
    prd_content = request.json.get('prd_content', None)
    prompt_id = request.json.get('prompt_id', None)
    chat_id = request.json.get('chat_id', None)
    if not prd_content:
        return error('需求为空')
    if not prompt_id:
        return error('prompt为空')
    prompt_row = db.session.query(Prompt).filter(Prompt.id == prompt_id).first()
    if not prompt_row:
        return error('未找到prompt模板')
    prompt_content = prompt_row.content
    complete_prompt = prompt_content.replace('{prd_content}', prd_content)
    return llm_common_ask(complete_prompt, chat_id)


@app.route('/api/generate_cases', methods=["POST"])
def generate_cases():
    prompt_id = request.json.get('prompt_id', None)
    prd_content = request.json.get('prd_content', None)
    chat_id = request.json.get('chat_id', None)
    if not prompt_id:
        return error('prompt为空')
    if not prd_content:
        return error('需求为空')
    prompt_row = db.session.query(Prompt).filter(Prompt.id == prompt_id).first()
    if not prompt_row:
        return error('未找到prompt模板')
    prompt_content = prompt_row.content
    complete_prompt = prompt_content.replace('{prd_content}', prd_content)
    temp = """

根据提供的需求文档，用例生成要求生成测试用例，要求尽可能覆盖需求文档中的各个场景。
        """
    complete_prompt = complete_prompt + temp
    return llm_common_ask(complete_prompt, chat_id)


@app.route('/api/generate_cases_by_title', methods=["POST"])
def generate_cases_by_title():
    prompt_id = request.json.get('prompt_id', None)
    prd_url = request.json.get('prd_url', None)
    prd_content = request.json.get('prd_content', None)
    chat_id = request.json.get('chat_id', None)
    titles = request.json.get('titles', None)
    if not prompt_id:
        return error('prompt为空')
    if not prd_url and not prd_content:
        return error('需求为空')
    if not titles:
        return error('至少填写一个用例标题')
    prompt_row = db.session.query(Prompt).filter(Prompt.id == prompt_id).first()
    if not prompt_row:
        return error('未找到prompt模板')
    if not prd_content:
        prd_content = '\n'.join([x.page_content for x in pdf_loader(prd_url)])
    prompt_content = prompt_row.content
    complete_prompt = prompt_content.replace('{prd_content}', prd_content)
    temp = """
    
根据以下提供的用例标题补充json格式的测试用例: 
{}
    """.format('\n'.join(f'{index + 1}、{title}' for index, title in enumerate(titles)))
    complete_prompt = complete_prompt + temp
    return llm_common_ask(complete_prompt, chat_id)


@app.route('/api/create_single_case_by_title', methods=["POST"])
def create_single_case_by_title():
    prompt_id = request.json.get('prompt_id', None)
    prd_url = request.json.get('prd_url', None)
    prd_content = request.json.get('prd_content', None)
    chat_id = request.json.get('chat_id', None)
    title = request.json.get('title', None)
    if not prd_url and not prd_content:
        return error('需求为空')
    if not title:
        return error('至少填写一个用例标题')
    if not prompt_id:
        prompt_row = db.session.query(Prompt).filter(Prompt.role == 8).first()
    else:
        prompt_row = db.session.query(Prompt).filter(Prompt.id == prompt_id).first()
    if not prompt_row:
        return error('未找到prompt模板')
    if not prd_content:
        prd_content = '\n'.join([x.page_content for x in pdf_loader(prd_url)])
    prompt_content = prompt_row.content
    complete_prompt = prompt_content.replace('{prd_content}', prd_content)

    complete_prompt = complete_prompt.replace('{case_title}', title)
    return llm_common_ask(complete_prompt, chat_id)


@app.route('/api/case/optimise', methods=["POST"])
def optimise_cases():
    prompt_id = request.json.get('prompt_id', None)
    cases = request.json.get('cases', None)
    demand = request.json.get('demand', None)
    chat_id = request.json.get('chat_id', None)
    if not cases:
        return error('优化用例为空')
    if not prompt_id:
        return error('prompt为空')
    if not demand:
        return error('优化要求为空')
    prompt_row = db.session.query(Prompt).filter(Prompt.id == prompt_id).first()
    if not prompt_row:
        return error('未找到prompt模板')
    prompt_content = prompt_row.content
    complete_prompt = prompt_content.replace('{cases}', json.dumps(cases, ensure_ascii=False))
    complete_prompt = complete_prompt.replace('{demand}', demand)
    return llm_common_ask(complete_prompt, chat_id)


@app.route('/api/save_cases', methods=["POST"])
def save_cases():
    case = request.json.get('case', None)
    cases = request.json.get('cases', None)
    project_id = request.json.get('project_id', None)
    sprint_id = request.json.get('sprint_id', None)
    node_id = request.json.get('node_id', 0)
    if not project_id:
        return error('项目id为空')
    if not sprint_id:
        return error('迭代id为空')
    if not case and not cases:
        return error('未找到用例')
    creator = request.cookies.get('username', None)
    if case:
        case_model = save_case_result(case, project_id, sprint_id, node_id, creator)
        db.session.add(case_model)
    if cases:
        for case in cases:
            case_model = save_case_result(case, project_id, sprint_id, node_id, creator)
            db.session.add(case_model)
    db.session.commit()
    return success()


@app.route('/api/save_optimised_cases', methods=["POST"])
def save_optimised_cases():
    cases = request.json.get('cases', None)
    project_id = request.json.get('project_id', None)
    sprint_id = request.json.get('sprint_id', None)
    node_id = request.json.get('node_id', 0)
    if not project_id:
        return error('没有project id')
    if not sprint_id:
        return error('没有sprint id')
    for case in cases:
        row = db.session.query(TestCase).filter(TestCase.id == case['id']).first()
        if not row:
            return error('未找到用例id:{}'.format(case['id']))
        rtn = save_optimised_case(row, case, project_id, sprint_id, node_id)
        if not rtn:
            return error('保存用例出错，请确认用例字段存在且正确')
    db.session.commit()
    return success()


@app.route('/api/case/edit', methods=["POST"])
def case_edit():
    case_id = request.json.get('case_id', None)
    case = db.session.query(TestCase).filter(TestCase.id == case_id).first()
    if not case:
        return error('未找到该数据')
    case.node_id = request.json.get('node_id')
    case.name = request.json.get('name')
    case.priority = request.json.get('priority')
    case.precondition = request.json.get('precondition')
    case.description = json.dumps(request.json.get('description'), ensure_ascii=False)
    db.session.commit()
    return success()


@app.route('/api/get_cases', methods=["POST"])
@require_login
def get_cases():
    project_id = request.json.get('project_id', None)
    sprint_id = request.json.get('sprint_id', None)
    node_id = request.json.get('node_id', None)
    query = db.session.query(TestCase, Sprint.name).join(Sprint, TestCase.sprint_id == Sprint.id).filter(
        TestCase.project_id == project_id)
    if sprint_id:
        query = query.filter(TestCase.sprint_id == sprint_id)
    if node_id:
        def get_offspring_nodes(id):
            rtn = [id]
            childs = db.session.query(Nodes).filter(Nodes.parent_id == id).all()
            if childs:
                for child in childs:
                    rtn.extend(get_offspring_nodes(child.id))
            return rtn

        nodes = get_offspring_nodes(node_id)
        query = query.filter(TestCase.node_id.in_(nodes))
    result = query.order_by(-TestCase.id).all()

    def get_review_status(case: TestCase):
        query = db.session.query(CaseReview).filter(CaseReview.case_id == case.id).order_by(-CaseReview.id).first()
        return query.result if query else None

    return success([{**case.as_dict(),
                     'sprint_name': sprint_name,
                     'description': json.loads(case.description),
                     'review_status': get_review_status(case),
                     } for case, sprint_name in result])


@app.route('/api/case_detail', methods=["POST"])
def get_case_detail():
    case_id = request.json.get('case_id', None)
    if not case_id:
        return error('没有case id')
    query = db.session.query(TestCase, Sprint.name).join(Sprint, TestCase.sprint_id == Sprint.id).filter(
        TestCase.id == case_id).first()
    if not query:
        return error('未搜索到用例')

    def get_review_status(case: TestCase):
        query = db.session.query(CaseReview).filter(CaseReview.case_id == case.id).order_by(-CaseReview.id).first()
        return query.result if query else None

    case, sprint_name = query
    return success({
        **case.as_dict(),
        'sprint_name': sprint_name,
        'description': json.loads(case.description),
        'review_status': get_review_status(case),
    })


@app.route('/api/delete_cases', methods=["POST"])
def delete_cases():
    case_id = request.json.get('case_id', None)
    if not case_id:
        error('未找到用例id')
    row = db.session.query(TestCase).filter(TestCase.id == case_id).first()
    db.session.delete(row)
    db.session.commit()
    return success()


@app.route('/api/cases/bunch_delete', methods=["POST"])
def case_bunch_delete():
    case_ids = request.json.get('case_ids', None)
    if not case_ids:
        return error('请输入用例id')
    rows = db.session.query(TestCase).filter(TestCase.id.in_(case_ids)).all()
    if not rows:
        return error('未找到用例')
    for row in rows:
        db.session.delete(row)
    db.session.commit()
    return success()


@app.route('/api/cases/bunch_move', methods=["POST"])
def case_bunch_move():
    case_ids = request.json.get('case_ids', None)
    target_node_id = request.json.get('target_node_id', None)
    if not case_ids or target_node_id == None:
        return error('请输入用例id和目标模块')
    rows = db.session.query(TestCase).filter(TestCase.id.in_(case_ids)).all()
    if not rows:
        return error('未找到用例')
    for row in rows:
        row.node_id = target_node_id
    db.session.commit()
    return success()


@app.route('/api/cases/bunch_copy', methods=["POST"])
def case_bunch_copy():
    case_ids = request.json.get('case_ids', None)
    target_node_id = request.json.get('target_node_id', None)
    if not case_ids or target_node_id == None:
        return error('请输入用例id和目标模块')
    rows = db.session.query(TestCase).filter(TestCase.id.in_(case_ids)).all()
    if not rows:
        return error('未找到用例')
    for row in rows:
        new_case = TestCase()
        new_case.name = 'Copy of ' + row.name
        new_case.priority = row.priority
        new_case.precondition = row.precondition
        new_case.description = row.description
        new_case.sprint_id = row.sprint_id
        new_case.project_id = row.project_id
        new_case.node_id = target_node_id
        new_case.creator = request.cookies.get('username', None)
        new_case.status = 1
        db.session.add(new_case)
    db.session.commit()
    return success()


def get_modules(node_id):
    def get_parent_node(id):
        node_info = db.session.query(Nodes).filter(Nodes.id == id).first()
        if not node_info:
            return ''
        if node_info.parent_id:
            return get_parent_node(node_info.parent_id) + '/' + node_info.name
        return node_info.name

    return get_parent_node(node_id)


def export_cases_to_excel(project_id, sprint_id, node_id, case_list, ms_node_path='迭代用例'):
    from openpyxl import Workbook
    from io import BytesIO
    query = db.session.query(TestCase, Sprint.name).join(Sprint, TestCase.sprint_id == Sprint.id).filter(
        TestCase.project_id == project_id)
    if case_list:  # 有指定用例编号列表，只导出该用例, sprint id和node id都不看了
        query = query.filter(TestCase.id.in_(case_list))
    else:
        if sprint_id:
            query = query.filter(TestCase.sprint_id == sprint_id)
        if node_id:
            def get_offspring_nodes(id):
                rtn = [id]
                childs = db.session.query(Nodes).filter(Nodes.parent_id == id).all()
                if childs:
                    for child in childs:
                        rtn.extend(get_offspring_nodes(child.id))
                return rtn

            nodes = get_offspring_nodes(node_id)
            query = query.filter(TestCase.node_id.in_(nodes))
    rows = query.order_by(-TestCase.id).all()
    case_num = query.count()
    wb = Workbook()
    ws = wb.active
    ws.append(['用例名称', '所属模块', '标签', '前置条件', '备注', '步骤描述', '预期结果', '编辑模式', '用例状态', '责任人',
               '用例等级', '用例阶段'])
    start_row = 2
    for row, sprint_name in rows:
        modules = get_modules(row.node_id)
        description = json.loads(row.description)
        need_cnt = len(description)
        if need_cnt > 1:
            for letter in ['A', 'B', 'C', 'D', 'E', 'H', 'I', 'J', 'K', 'L']:  # 合并单元格
                ws.merge_cells("{letter}{start}:{letter}{end}".format(
                    letter=letter, start=start_row, end=start_row + need_cnt - 1))
        ws['A{}'.format(start_row)] = row.name
        ws['D{}'.format(start_row)] = row.precondition
        for index, des in enumerate(description):
            if not des:
                break
            ws['F{}'.format(start_row + index)] = des['steps'] if 'steps' in des else ''
            ws['G{}'.format(start_row + index)] = des['expect_result'] if 'expect_result' in des else ''
        ws['B{}'.format(start_row)] = ms_node_path + '/' + sprint_name + '/' + modules  # 所属模块
        ws['I{}'.format(start_row)] = '未开始'  # 用例状态固定未开始
        ws['J{}'.format(start_row)] = 'AI大模型'  # 责任人固定AI大模型
        ws['K{}'.format(start_row)] = row.priority
        ws['L{}'.format(start_row)] = '手工'  # 用例阶段固定手工

        start_row += need_cnt

    file_obj = BytesIO()
    wb.save(file_obj)
    # wb.save('text.xlsx')
    file_obj.seek(0)
    return case_num, file_obj


@app.route('/api/export_cases', methods=["POST"])
def export_cases():
    project_id = request.json.get('project_id', None)
    sprint_id = request.json.get('sprint_id', None)
    node_id = request.json.get('node_id', None)
    case_list = request.json.get('case_list', None)
    total, file_obj = export_cases_to_excel(project_id, sprint_id, node_id, case_list)
    return send_file(file_obj, download_name='cases-{}.xlsx'.format(datetime.datetime.now().strftime('%Y%m%d%H%M%S')))


def get_technical_center_ms_project(sprint_name, sub_projects):
    import re
    rtn = re.search(r'(.*?)[vV]\d+\.\d+\.\d+', sprint_name)
    if not rtn:
        return None
    rows = [row for row in sub_projects if rtn.groups()[0].strip() == row.name]
    if not rows:
        return None
    return rows[0].ms_project


@app.route('/api/import_cases', methods=["POST"])
def import_cases():
    project_id = request.json.get('project_id', None)
    sprint_id = request.json.get('sprint_id', None)
    node_id = request.json.get('node_id', None)
    case_list = request.json.get('case_list', None)
    ms_node_path = request.json.get('ms_node_path', None)
    if not project_id:
        return error('项目未找到')
    project_row = db.session.query(Projects).filter(Projects.id == project_id).first()
    if project_row.name == '技术中台':  # 技术中台从子项目中找配置
        sub_project_rows = db.session.query(Projects).filter(Projects.parent_id == project_row.id).all()
        sprint_row = db.session.query(Sprint).filter(Sprint.id == sprint_id).first()
        if not sprint_row:
            return error('未找到该sprint:{}'.format(sprint_id))
        ms_project = get_technical_center_ms_project(sprint_row.name, sub_project_rows)
    else:
        ms_project = project_row.ms_project
    if not ms_project:
        return error('未找到配置的metersphere项目')
    case_num, file_obj = export_cases_to_excel(project_id, sprint_id, node_id, case_list, ms_node_path)
    file_name = 'cases-{}.xlsx'.format(datetime.datetime.now().strftime('%Y%m%d%H%M%S'))
    ms_utils = MSUtils(project_name=ms_project)
    rtn = ms_utils.import_excel_case(file_name, file_obj)
    if rtn['success']:
        if rtn['data']['errList']:
            return success({
                'total': case_num,
                'failed': rtn['data']['errList']
            })
        return success({
            'total': case_num,
            'failed': []
        })
    else:
        return error('导入失败，message:{}'.format(rtn['message']))


def get_node_cases(node_id, sprint_id):
    rtn = []
    # 获取用例标题数据
    cases = db.session.query(TestCase).filter(TestCase.node_id == int(node_id), TestCase.sprint_id == sprint_id).all()
    rtn = rtn + [
        {"id": 'case-{}'.format(x.id), "parentid": 'module-{}'.format(node_id), "isroot": False, "topic": x.name,
         "expanded": False,
         "data": {'type': 'case', 'priority': x.priority, "topic": x.name}} for x in cases]
    # 获取用例前置条件、步骤和结果
    for case in cases:
        if case.precondition:
            pre = {"id": 'casePre-{}'.format(case.id), "parentid": 'case-{}'.format(case.id),
                   "isroot": False, "topic": case.precondition,
                   "data": {'type': 'casePre', "topic": case.precondition}}
            rtn.append(pre)
        for index, des in enumerate(json.loads(case.description)):
            if 'steps' in des:
                step = {"id": 'caseStep-{}-{}'.format(case.id, index), "parentid": 'case-{}'.format(case.id),
                        "isroot": False, "topic": des['steps'],
                        "data": {'type': 'caseStep', "topic": des['steps']}}
                rtn.append(step)
                if 'expect_result' in des:
                    result = {"id": 'caseResult-{}-{}'.format(case.id, index), "parentid": step['id'], "isroot": False,
                              "topic": des['expect_result'], "expanded": False,
                              "data": {'type': 'caseResult', "topic": des['expect_result']}}
                    rtn.append(result)
    return rtn


@app.route('/api/case/mind/init', methods=["POST"])
@require_login
def case_mind_init():
    nodes_data = request.json.get('nodes_data', None)
    sprint_id = request.json.get('sprint_id', None)
    node_id = request.json.get('node_id', None)

    def filter_target_nodes_data(nodes):
        for row in nodes:
            if row['key'] == node_id:
                return [row]
            elif row['children']:
                rtn = filter_target_nodes_data(row['children'])
                if rtn:
                    return rtn
    if node_id:
        nodes_data = filter_target_nodes_data(nodes_data)
    if not nodes_data:
        return success([])
    mind_list = []

    def get_mind_data(nodes, parent_id=None):
        for node in nodes:
            node_id = node['key']
            mind_list.append(
                {"id": 'module-{}'.format(node_id), "isroot": False if parent_id else True, 'parentid': parent_id,
                 "topic": node['title'], "data": {'type': 'module', "topic": node['title']}, 'expanded': False})
            # 获取用例信息
            mind_list.extend(get_node_cases(node_id, sprint_id))
            if node['children']:
                get_mind_data(node['children'], 'module-{}'.format(node_id))

    get_mind_data(nodes_data)
    mind_list[0]['isroot'] = True
    return success(mind_list)


def concat_case_description(nodes):
    rtn = []
    for node in nodes:
        if node['data']['type'] != 'casePre':
            steps = node['data']['topic']
            expect_result = node['children'][0]['data']['topic'] if 'children' in node else ''
            rtn.append({"steps": steps, 'expect_result': expect_result})
    return json.dumps(rtn, ensure_ascii=False)


def find_precondition(nodes):
    pre_rows = [x for x in nodes if x['data']['type'] == 'casePre']
    if not pre_rows:
        return ''
    return pre_rows[0]['data']['topic']


@app.route('/api/case/mind/save', methods=["POST"])
@require_login
def case_mind_save():
    node_tree = request.json.get('node_tree', None)
    removed_ids = request.json.get('removed_ids', None)
    sprint_id = request.json.get('sprint_id', None)
    project_id = request.json.get('project_id', None)

    def save_change(nodes, parent_id):
        for node in nodes:
            if 'data' not in node or 'type' not in node['data']:
                return error('存在未标记的用例')
            node_type = node['data']['type']
            if node_type == 'module':  # 处理模块
                # 处理新增模块逻辑
                if '-' not in node['id']:  # 表示编辑过的
                    node_instance = Nodes()
                    node_instance.name = node['data']['topic']
                    node_instance.parent_id = None if not parent_id else parent_id
                    node_instance.sprint_id = sprint_id
                    node_instance.project_id = project_id
                    db.session.add(node_instance)
                    db.session.commit()
                    node_id = node_instance.id
                elif 'is_edit' in node and node['is_edit']:
                    node_id = node['id'].split('-')[1]
                    node_instance = db.session.query(Nodes).filter(Nodes.id == node_id).first()
                    if not node_instance:
                        return error('找不到node节点:{}'.format(node_id))
                    node_instance.name = node['data']['topic']
                else:
                    node_id = node['id'].split('-')[1]
                if 'children' in node:
                    save_change(node['children'], node_id)
            elif node_type == 'case':  # 处理用例
                # 新增用例的情况
                if '-' not in node['id']:
                    case = TestCase()
                    case.name = node['data']['topic']
                    case.priority = node['data']['priority']
                    case.precondition = find_precondition(node['children']) if 'children' in node else ''
                    case.description = concat_case_description(node['children']) if 'children' in node else '[]'
                    case.sprint_id = sprint_id
                    case.project_id = project_id
                    case.node_id = parent_id
                    case.creator = request.cookies.get('username', None)
                    case.status = 1
                    db.session.add(case)
                elif 'is_edit' in node and node['is_edit']:  # 编辑用例情况
                    case_id = node['id'].split('-')[1]
                    case = db.session.query(TestCase).filter(TestCase.id == int(case_id)).first()
                    if not case:
                        return error('找不到编辑的case:{}'.format(case_id))
                    case.name = node['data']['topic']
                    case.priority = node['data']['priority']
                    case.precondition = find_precondition(node['children']) if 'children' in node else ''
                    case.description = concat_case_description(node['children']) if 'children' in node else '[]'
                    case.sprint_id = sprint_id
                    case.project_id = project_id
                    case.node_id = parent_id
                    case.creator = request.cookies.get('username', None)
                    case.status = 1
        return success()

    rtn = save_change(
        node_tree['children'] if 'children' in node_tree else [],
        None if not node_tree['id'].split('-')[1] else int(node_tree['id'].split('-')[1])
                      )
    if not rtn['success']:
        return rtn
    # 数据库删除被删除的节点
    if removed_ids:
        for removed_id in removed_ids:
            node_type, node_id = removed_id.split('-')[0], removed_id.split('-')[1]
            if node_type == 'module':
                row = db.session.query(Nodes).filter(Nodes.id == int(node_id)).first()
                if row:
                    db.session.delete(row)
            if node_type == 'case':
                row = db.session.query(TestCase).filter(TestCase.id == int(node_id)).first()
                if row:
                    db.session.delete(row)
    db.session.commit()
    return success()
